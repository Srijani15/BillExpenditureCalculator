import tkinter as tk
from tkinter import filedialog, messagebox
import pytesseract
import re
import cv2
from PIL import Image, ImageTk
import threading
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
import pyttsx3
import pygame
from pygame import mixer

engine = pyttsx3.init()  # Text to speech conversion

# Initialize Pygame and load the sound files
pygame.mixer.init()
click_sound = pygame.mixer.Sound(r"C:/Users/srija/Downloads/click_button.mp3")

# Function to play click sound
def play_click_sound():
    pygame.mixer.Sound.play(click_sound)

# Function to extract text from image using pytesseract
def extract_text_from_image(image_path):
    try:
        text = pytesseract.image_to_string(image_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to extract text from image: {e}")
        return ""
    return text

# Function to analyze bill and extract total amount
def analyze_bill(text):
    total_amount_match = re.search(r"^Total\s*[^0-9,.]*([₹]*[\d,]+\.\d{2}|\d+)", text, re.IGNORECASE | re.MULTILINE)
    if total_amount_match:
        total_amount = total_amount_match.group(1).replace(',', '')
    else:
        total_amount = None
    
    #amounts = re.findall(r'\d*[$|,]\d+\.\d{2}', text)

    return total_amount

# Function to save extracted total amounts to an Excel file
def save_to_excel(image_path, total_amount):
    if not os.path.exists("bill_data.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["BILL NUMBER", "  TOTAL AMOUNT  " ])
    else:
        wb = openpyxl.load_workbook("bill_data.xlsx")
        ws = wb.active

    ws.append([image_path, total_amount ])

     # Apply styling to the cells
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=12)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    wb.save("bill_data.xlsx")

# Function to validate the total amount with user input
def validate_total(total_amount):
    if total_amount is None:
        messagebox.showinfo("Validation", "Total amount not found in the bill.")
        return False
    user_input = messagebox.askyesno("Validation", f"The total amount extracted is {total_amount}. Is this correct?")
    return user_input

# Function to show the validation result
def show_result_page(success, text, image_path , total_amount):
    # Clear the canvas content
    for widget in canvas.winfo_children():
        widget.destroy()

    # Set background color based on success or failure
    bg_color = "#d4edda" if success else "#f8d7da"
    canvas.configure(bg=bg_color)

    # Display the appropriate message
    message = "Validation Successful!" if success else "Validation Failed!"
    result_text_label = tk.Label(canvas, text=message, font=("Helvetica", 16), bg=bg_color)
    result_text_label.place(relx=0.77, rely=0.8, anchor=tk.CENTER)

    icon_file = r"C:/Users/srija/Downloads/check_mark.png" if success else r"C:/Users/srija/Downloads/wrong_mark.png"
    try:
        icon = Image.open(icon_file)
        icon = icon.resize((80, 80))
        icon_tk = ImageTk.PhotoImage(icon)
        icon_label = tk.Label(canvas, image=icon_tk, bg=bg_color)
        icon_label.image = icon_tk
        icon_label.place(relx=0.9, rely=0.8, anchor=tk.CENTER)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load icon: {e}")

    try:
        input_image = Image.open(image_path)
        input_image = input_image.resize((400, 400))
        input_image_tk = ImageTk.PhotoImage(input_image)
        input_image_label = tk.Label(canvas, image=input_image_tk, bg=bg_color)
        input_image_label.image = input_image_tk
        input_image_label.place(relx=0.8, rely=0.4, anchor=tk.CENTER)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load input image: {e}")

    # Creating a frame for the text, heading, and scrollbar
    text_frame = tk.Frame(canvas, bg=bg_color)
    text_frame.place(relx=0.3, rely=0.45, anchor=tk.CENTER, relwidth=0.4, relheight=0.8)

    # Creating a heading label for the text widget
    heading_label = tk.Label(text_frame, text="Extracted Text", font=("Rockwell", 14, "bold"), bg=bg_color)
    heading_label.pack(side="top", fill="x")

    # Creating a text widget and scrollbar
    text_scrollbar = tk.Scrollbar(text_frame)
    text_widget = tk.Text(text_frame, wrap="word", yscrollcommand=text_scrollbar.set, bg=bg_color, font=("Helvetica", 12))
    text_scrollbar.config(command=text_widget.yview)

    # Insert the extracted text
    text_widget.insert(tk.END, text)
    text_widget.configure(state="disabled")

    # Pack the text widget and scrollbar
    text_scrollbar.pack(side="right", fill="y")
    text_widget.pack(side="left", fill="both", expand=True)

    close_button = tk.Button(canvas, text="CLOSE", command=on_close, font=("Rockwell", 14), bg="#4CAF50", fg="white")
    close_button.place(relx=0.78, rely=0.95, anchor=tk.CENTER)

    # Speak the result
    engine.say(message)
    engine.runAndWait()

    # Create a menu bar
    menu_bar = tk.Menu(root)
    root.config(menu=menu_bar)

    # Add 'File' menu
    file_menu = tk.Menu(menu_bar, tearoff=0)
    menu_bar.add_cascade(label="File", menu=file_menu)
    file_menu.add_command(label="Open Image", command=lambda: [play_click_sound(), display_input_image(image_path)])
    file_menu.add_command(label="Show Extracted Text", command=lambda: [play_click_sound(), messagebox.showinfo("Extracted Text", text)])
    file_menu.add_separator()
    file_menu.add_command(label="Exit", command=on_close)

    # Add 'Report' menu
    report_menu = tk.Menu(menu_bar, tearoff=0)
    menu_bar.add_cascade(label="Report", menu=report_menu)
    report_menu.add_command(label="Prepare Report", command=lambda: [play_click_sound(), prepare_report()])

# Function to handle the scan button click
def scan_image():
    file_path = filedialog.askopenfilename(title="Select an image", filetypes=[("Image files", ".png;.jpg;*.jpeg")])
    if file_path:
        display_input_image(file_path)

# Function to display the input image for confirmation
def display_input_image(image_path):
    for widget in canvas.winfo_children():
        widget.destroy()

    try:
        input_image = Image.open(image_path)
        input_image = input_image.resize((500, 500))
        input_image_tk = ImageTk.PhotoImage(input_image)
        input_image_label = tk.Label(canvas, image=input_image_tk)
        input_image_label.image = input_image_tk
        input_image_label.place(relx=0.5, rely=0.4, anchor=tk.CENTER)

        confirm_button = tk.Button(canvas, text="CONFIRM", command=lambda: [play_click_sound(), proceed_to_validation(image_path)], font=("Rockwell", 14), bg="#4CAF50", fg="white")
        confirm_button.place(relx=0.5, rely=0.83, anchor=tk.CENTER)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to display input image: {e}")

index = 1
# Function to proceed to validation after clicking confirm
def proceed_to_validation(image_path):
    global index
    text = extract_text_from_image(image_path)
    if text:
        total_amount= analyze_bill(text)
        #amounts = analyze_bill(text)
        if total_amount:
            if validate_total(total_amount):
                save_to_excel(f"    Bill - {index}    ", total_amount)
                index += 1
                show_result_page(True, text, image_path, total_amount)
            else:
                show_result_page(False, text, image_path, total_amount)
        else:
            show_result_page(False, text, image_path, "Not Found")
    else:
        messagebox.showinfo("Error", "Failed to extract text from the image.")

# Function to prepare and display the report
def prepare_report():
    if not os.path.exists("bill_data.xlsx"):
        messagebox.showinfo("Report", "No data available to prepare the report.")
        return

    wb = openpyxl.load_workbook("bill_data.xlsx")
    ws = wb.active

    # Calculate the total expenditure
    total_expenditure = 0
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        price = row[1]
        try:
            if row[0]== "Total Expenditure" :
                continue
            else:
                total_expenditure += float(price)
        except ValueError:
            continue

    # Append total expenditure to the last row
    ws.append(["Total Expenditure", total_expenditure])
    
    # Apply styling to the total expenditure row
    last_row = ws.max_row
    for col in range(1, ws.max_column + 1):
        ws.cell(row=last_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=last_row, column=col).font = Font(bold=True)
        ws.cell(row=last_row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    wb.save("bill_data.xlsx")

    messagebox.showinfo("Report", f"Total Expenditure: {total_expenditure}")

    # Open the Excel file
    os.system("start EXCEL.EXE bill_data.xlsx")

# Function to handle the close event
def on_close():
    file_path = "bill_data.xlsx"
    if os.path.exists(file_path):
        os.remove(file_path)
    root.destroy()

mixer.init()

# Function to display the animated video in the GUI
def show_video():

    # Load and play music
    mixer.music.load(r"C:/Users/srija/Downloads/joyride-jamboree-206911.mp3")
    mixer.music.play(-1)    # -1 ensures the music loops indefinitely

    cap = cv2.VideoCapture(r"C:/Users/srija/Videos/Captures/input_video.mp4")
    if not cap.isOpened():
        messagebox.showerror("Error", "Cannot open the video file")
        return

    def update_frame():
        ret, frame = cap.read()
        if ret:
            cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(cv2image)
            imgtk = ImageTk.PhotoImage(image=img)
            video_label.imgtk = imgtk
            video_label.configure(image=imgtk)
            video_label.after(10, update_frame)
        else:
            cap.release()
            mixer.music.stop()   # Stop music when the video ends
            show_scan_buttons()

    update_frame()

# Function to show the scan button
def show_scan_buttons():
    scan_button.place(relx=0.43, rely=0.92, anchor=tk.CENTER)
    scan_button_realtime.place(relx=0.6, rely=0.92, anchor=tk.CENTER)


def start_real_time_scan():
    # Try using DirectShow backend for higher quality if available
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    
    # If DirectShow backend fails, try using V4L2 backend
    if not cap.isOpened():
        cap = cv2.VideoCapture(0, cv2.CAP_V4L2)
        
    if not cap.isOpened():
        messagebox.showerror("Error", "Cannot access the camera")
        return

    # Set camera properties for better quality
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 800)  # Set width
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 600)  # Set height
    cap.set(cv2.CAP_PROP_FPS, 30)  # Set frame rate

    ret, frame = cap.read()
    if ret:
        display_camera_feed(frame, cap)
    else:
        messagebox.showerror("Error", "Failed to capture the image")


def display_camera_feed(frame, cap):
    def save_and_proceed():
        ret, frame = cap.read()
        if ret:
            cv2.imwrite("real_time_scan.png", frame)
            cap.release()
            cam_window.destroy()
            proceed_to_validation("real_time_scan.png")
        else:
            messagebox.showerror("Error", "Failed to capture the image")

    cam_window = tk.Toplevel(root)
    cam_window.title("Real-Time Scan")

    camera_label = tk.Label(cam_window)
    camera_label.pack()

    confirm_button = tk.Button(cam_window, text="CONFIRM", command=save_and_proceed, font=("Rockwell", 14), bg="#4CAF50", fg="white")
    confirm_button.pack()

    def update_camera_frame():
        ret, frame = cap.read()
        if ret:
            cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(cv2image)
            imgtk = ImageTk.PhotoImage(image=img)
            camera_label.imgtk = imgtk
            camera_label.configure(image=imgtk)
            camera_label.after(10, update_camera_frame)
        else:
            cap.release()
            cam_window.destroy()

    update_camera_frame()


#Saving the image that camera captured
def save_image(frame):
    cv2.imwrite("real_time_scan.png", frame)
    proceed_to_validation("real_time_scan.png")

# Create the main application window
root = tk.Tk()
root.title("BILL SCANNER")
root.geometry("1200x800")

canvas = tk.Canvas(root, width=1200, height=800)
canvas.pack()

video_label = tk.Label(canvas)
video_label.place(relx=0.5, rely=0.45, anchor=tk.CENTER)

scan_button = tk.Button(root, text="SCAN", command=lambda: [play_click_sound(), scan_image()], font=("Rockwell", 16), bg="skyblue", fg="black", width=10, height=2)
scan_button_realtime = tk.Button(root, text="REAL-TIME SCAN", command=lambda: [play_click_sound(), start_real_time_scan()], font=("Rockwell", 16), bg="skyblue", fg="black", width=15, height=2)

video_thread = threading.Thread(target=show_video)
video_thread.start()

root.protocol("WM_DELETE_WINDOW", on_close)
root.attributes('-fullscreen',True)
root.mainloop()
